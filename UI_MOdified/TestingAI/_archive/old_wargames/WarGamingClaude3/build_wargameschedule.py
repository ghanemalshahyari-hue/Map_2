"""
build_wargameschedule.py
========================
Builds wargameschedule.xlsx in the long-format matrix the user specified:

  Col 1: Phase number (step 0..16, phase label, time)
  For each phase, MULTIPLE ROWS — one per component (Land/Navy/Air/SOF/...).
  Col 2: Component
  Col 3: Red action
  Col 4: Blue reaction
  Col 5: Red reaction to reaction (counter)
  Col 6: Goal / reasoning (why)
  Col 7: Effect of this action

Phase number column is merged vertically across the component rows of each phase.
Color-coded by component for quick scanning.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
STATE = json.load(open(ROOT / "qa" / "simulation_full_state.json"))
SNAPS = STATE["snapshots"]

# ------------------------------------------------------------------
# Style helpers
# ------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
HEADER_FONT = Font(color="ffffff", bold=True, size=11)
PHASE_FILL  = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")
PHASE_FONT  = Font(bold=True, size=11, color="78350f")

COMPONENT_COLORS = {
    "استراتيجي (SSM)":      "f5d0fe",  # purple
    "بحري (Navy)":          "bae6fd",  # sky
    "جوي (Air)":            "fed7aa",  # orange
    "حرب ألغام (Mines)":    "fef08a",  # yellow
    "USV / UAV":            "fecaca",  # rose
    "عمليات خاصة (SOF)":    "ddd6fe",  # violet
    "بري (Land)":           "bbf7d0",  # green
    "حرب إلكترونية (EW)":   "e5e7eb",  # gray
}

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(border_style="thin", color="9ca3af")
THICK = Side(border_style="medium", color="1e3a8a")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PHASE_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

# ------------------------------------------------------------------
# Components and the geojson arrays they map to
# ------------------------------------------------------------------
COMPONENTS = [
    ("استراتيجي (SSM)",   "strategic_strikes"),
    ("بحري (Navy)",       "maritime_actions"),
    ("جوي (Air)",          "air_actions"),
    ("حرب ألغام (Mines)", "mine_warfare"),
    ("USV / UAV",          "uav_usv_swarms"),
    ("عمليات خاصة (SOF)", "sof_actions"),
    ("بري (Land)",         "land_actions"),
    ("حرب إلكترونية (EW)","ew_actions"),
]

# ------------------------------------------------------------------
# Helpers to split actions by side and extract fields
# ------------------------------------------------------------------
def is_red(actor):
    a = (actor or "").strip().upper()
    return a.startswith("R-") or a.startswith("RED")

def is_blue(actor):
    a = (actor or "").strip().upper()
    return a.startswith("B-") or a.startswith("BLUE")

def fmt_one(e, prefix=""):
    """Format a single action dict as a readable line."""
    actor = e.get("actor","?")
    what = e.get("what","")
    extras = []
    for k in ("targets","sorties","missiles","launched","hits","kills",
              "surviving","intercepted_by_blue_helos","intercepted_by_CRAM",
              "intercepted_by_AD","intercepted_by_CIWS","method","clearance_rate",
              "remaining","helos","casualties_inflicted","casualties_landing",
              "losses_at_drop","defender_alert","red_losses","blue_losses",
              "warheads_effective","blue_armor_kills","red_helo_losses",
              "red_intercept_losses","red_helo_kills","blue_intercept_losses",
              "MANPADS_kills","red_helo_total_losses","red_air_intercept_losses",
              "intensity"):
        v = e.get(k)
        if v is not None and v != "":
            extras.append(f"{k}={v}")
    extras_s = f"  ({', '.join(extras)})" if extras else ""
    return f"{prefix}{actor}: {what}{extras_s}"

def entries_for_side(arr, side_filter):
    """Return list of full entry dicts whose actor matches the side filter."""
    out = []
    for e in arr or []:
        actor = e.get("actor","")
        if side_filter == "R" and not is_red(actor): continue
        if side_filter == "B" and not is_blue(actor): continue
        out.append(e)
    return out

def fmt_why(e):
    """Extract the 'why' for a single entry."""
    parts = []
    for k in ("why","intended_effect","doctrine","calibration"):
        v = e.get(k)
        if v: parts.append(v if k in ("why","intended_effect") else f"({k}: {v})")
    return " — ".join(parts) if parts else "—"

def fmt_effect(e):
    """Extract the 'effect' for a single entry, including numeric metrics."""
    parts = []
    for k in ("effect","effectiveness","outcome"):
        v = e.get(k)
        if v: parts.append(v)
    nums = []
    for k in ("hits","kills","losses","sorties","missiles","launched","warheads_effective",
              "intercepted_by_blue_helos","intercepted_by_CRAM","intercepted_by_AD","intercepted_by_CIWS",
              "red_losses","blue_losses","casualties_inflicted","casualties_landing","losses_at_drop",
              "blue_armor_kills","red_helo_losses","red_helo_kills","red_helo_total_losses",
              "MANPADS_kills","red_intercept_losses","blue_intercept_losses","red_AD_kills",
              "red_air_intercept_losses","intensity","remaining","surviving"):
        v = e.get(k)
        if v is not None and v != "":
            nums.append(f"{k}={v}")
    if nums:
        parts.append(", ".join(nums))
    return " — ".join(parts) if parts else "—"

def list_block(entries, key_fn):
    """Apply key_fn to each entry, join with newlines."""
    if not entries: return "—"
    out = []
    for e in entries:
        actor = e.get("actor","?")
        val = key_fn(e)
        out.append(f"{actor}: {val}" if val != "—" else f"{actor}: —")
    return "\n".join(out)

def actions_list(entries):
    if not entries: return "—"
    return "\n".join(fmt_one(e) for e in entries)

# ------------------------------------------------------------------
# Logistics + ammunition tally per (component, side)
# ------------------------------------------------------------------
COMPONENT_DOMAIN_KEY = {
    "استراتيجي (SSM)":   "strategic",
    "بحري (Navy)":       "naval",
    "جوي (Air)":          "air",
    "حرب ألغام (Mines)": "mines",      # virtual: filtered by type
    "USV / UAV":          "usv_uav",    # virtual: filtered by type
    "عمليات خاصة (SOF)": "sof",
    "بري (Land)":        "ground",     # excludes SAMs which we route to AD
    "حرب إلكترونية (EW)":"ew",          # virtual: type=ew_bn
}

def units_for_component(snap_units, comp_name):
    """Filter the per-step unit list to those that belong to this component."""
    key = COMPONENT_DOMAIN_KEY[comp_name]
    out = []
    for u in snap_units:
        d = u.get("domain","")
        t = u.get("type","")
        if key == "strategic" and d == "strategic":
            out.append(u)
        elif key == "naval" and d == "naval":
            # exclude USV / mine inventory virtual units from main naval tally
            if "usv" in t.lower() or "mine" in t.lower():
                continue
            out.append(u)
        elif key == "air" and d == "air":
            # exclude attack helos handled by ground? Actually include all air units
            out.append(u)
        elif key == "mines":
            # mine layers, mine sweepers, mine inventory
            if "mine" in t.lower():
                out.append(u)
        elif key == "usv_uav":
            if "usv" in t.lower() or "uav" in t.lower():
                out.append(u)
        elif key == "sof" and d == "sof":
            out.append(u)
        elif key == "ground" and d == "ground":
            if "ew_bn" in t or t.startswith("sam") or "_ad" in t.lower():
                continue
            out.append(u)
        elif key == "ew" and "ew" in t.lower():
            out.append(u)
    return out

def inventory_str(units):
    """Build a concise inventory line for a list of units.
    Sums hulls / airframes / strength. Returns '—' if empty.
    """
    if not units: return "—"
    n_alive = sum(1 for u in units if not u["destroyed"])
    n_dead  = sum(1 for u in units if u["destroyed"])
    total_hulls = sum(u.get("hulls_remaining",0) or 0 for u in units if u.get("hulls_remaining") is not None and not u["destroyed"])
    total_airframes = sum(u.get("airframes",0) or 0 for u in units if u.get("airframes") is not None and not u["destroyed"])
    total_strength = sum(round(u.get("strength",0) or 0, 1) for u in units if not u["destroyed"])
    parts = [f"وحدات حية: {n_alive}/{n_alive+n_dead}"]
    if total_hulls > 0:
        parts.append(f"سفن: {int(total_hulls)}")
    if total_airframes > 0:
        parts.append(f"طائرات: {int(total_airframes)}")
    parts.append(f"قوة فعّالة: {total_strength:.1f}")
    return " | ".join(parts)

def cum_losses_for_comp(state, comp_name, side, step_inclusive):
    """Count cumulative losses of a side in this component up to and including step_inclusive."""
    key = COMPONENT_DOMAIN_KEY[comp_name]
    total = 0
    destroyed_count = 0
    partial_count = 0
    for snap in state["snapshots"][:step_inclusive+1]:
        for l in snap["losses_this_step"][side]:
            d = l.get("domain","")
            t = l.get("type","")
            match = False
            if key == "strategic" and d == "strategic": match = True
            elif key == "naval" and d == "naval" and "usv" not in t.lower() and "mine" not in t.lower(): match = True
            elif key == "air" and d == "air": match = True
            elif key == "mines" and "mine" in t.lower(): match = True
            elif key == "usv_uav" and ("usv" in t.lower() or "uav" in t.lower()): match = True
            elif key == "sof" and d == "sof": match = True
            elif key == "ground" and d == "ground" and "ew" not in t.lower() and "sam" not in t.lower(): match = True
            elif key == "ew" and "ew" in t.lower(): match = True
            if match:
                if l.get("partial"): partial_count += 1
                else: destroyed_count += 1
                total += 1
    if total == 0: return "0"
    return f"تراكمي: {total} (مدمر: {destroyed_count} | جزئي: {partial_count})"

AMMO_BEARING_TYPES = {
    "ssm_brigade":   "صاروخ أرض/أرض",
    "sam_s300":      "صاروخ S-300",
    "sam_hawk":      "صاروخ هوك",
    "sam_2":         "صاروخ SAM-2",
    "sam_15":        "صاروخ SAM-15",
    "aaa":           "ذخيرة 35مم",
    "ad_brigade":    "صواريخ دفاع جوي (لوائي)",
    "ad_bn":         "صواريخ دفاع جوي (كتيبة)",
}

def ammo_str(units):
    """Build ammo line. Only types in AMMO_BEARING_TYPES count.
    Returns '—' if no ammo-bearing units in this component.
    """
    if not units: return "—"
    ammo_lines = []
    for u in units:
        if u["destroyed"]: continue
        t = u.get("type","")
        if t not in AMMO_BEARING_TYPES: continue
        mag = u.get("magazine")
        if mag is None: continue
        label = u.get("name", u["uid"])
        kind = AMMO_BEARING_TYPES[t]
        ammo_lines.append(f"{label} [{kind}]: {max(0, int(mag))}")
    return "\n".join(ammo_lines) if ammo_lines else "—"

def join_lines(lst):
    return "\n".join(lst) if lst else "—"

def domain_keyword_for_component(comp_name):
    """Map component name to a keyword for filtering counter-reactions text."""
    return {
        "استراتيجي (SSM)":"SSM",
        "بحري (Navy)":"بحر",
        "جوي (Air)":"جو",
        "حرب ألغام (Mines)":"لغم",
        "USV / UAV":"USV",
        "عمليات خاصة (SOF)":"SOF",
        "بري (Land)":"بر",
        "حرب إلكترونية (EW)":"EW",
    }.get(comp_name, "")

def collect_red_counter(snap, comp_name):
    """Pull red counter-reactions whose actor or text matches this component. Returns entry dicts."""
    out = []
    patterns = {
        "استراتيجي (SSM)":   ["SSM","صاروخ"],
        "بحري (Navy)":       ["NAV","fleet","سفينة","غواصة","fac","cor","dd","ff","lstm","lcu"],
        "جوي (Air)":          ["AB-","AEW","HELO","CAP","TKR","UAV","سرب","طلعة","طيران","cas","sortie"],
        "حرب ألغام (Mines)": ["MIN","MSW","لغم","mine"],
        "USV / UAV":         ["USV","UAV","زورق مفخخ","drone"],
        "عمليات خاصة (SOF)":["SOF","21SOF","80SOF","خاص"],
        "بري (Land)":        ["MID","VAN","ARTY","AT","ENG","REC","INF","ARMOR","TANK","BDE","مدرع","مدفع","الفرقة","لواء","كتيبة"],
        "حرب إلكترونية (EW)":["EW","405","505","تشويش"],
    }
    keys = patterns.get(comp_name, [])
    for e in snap.get("red_counter_reactions", []) or []:
        actor = e.get("actor",""); what = e.get("what","")
        blob = f"{actor} {what}".lower()
        if any(k.lower() in blob for k in keys):
            out.append(e)
    return out

# ------------------------------------------------------------------
# Build workbook
# ------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "WarGame Schedule"
ws.sheet_view.rightToLeft = True

# Header row — 7 main columns (red action, why, blue reaction, why, combined effect) + 6 logistics/ammo tally
headers = [
    ("الطور / الزمن", 22),
    ("المكوّن", 22),
    ("عمل الأحمر", 55),
    ("لماذا (الأحمر)", 45),
    ("رد فعل الأزرق", 55),
    ("لماذا (الأزرق)", 45),
    ("الأثر / الخلاصة الإجمالية", 60),
    ("قوة/عتاد الأحمر", 30),
    ("خسائر الأحمر التراكمية", 28),
    ("قوة/عتاد الأزرق", 30),
    ("خسائر الأزرق التراكمية", 28),
    ("ذخيرة الأحمر", 24),
    ("ذخيرة الأزرق", 24),
]
for i, (txt, w) in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=txt)
    ws.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[1].height = 38

current_row = 2

for snap in SNAPS:
    step = snap["step"]
    phase_label = f"الخطوة {step}\n{snap['time']}\n{snap['phase']}\nFR={snap['force_ratio_local']}:1\n{snap['step_advantage']}"
    rows_for_phase = []

    step_idx = snap["step"]
    red_snap_units  = snap["red_snapshot"]
    blue_snap_units = snap["blue_snapshot"]

    def combined_effects(red_entries, blue_entries):
        """Merge both sides' effects into a single conclusion block."""
        out = []
        for e in red_entries:
            actor = e.get("actor","?")
            ef = fmt_effect(e)
            if ef and ef != "—":
                out.append(f"[أحمر] {actor}: {ef}")
        for e in blue_entries:
            actor = e.get("actor","?")
            ef = fmt_effect(e)
            if ef and ef != "—":
                out.append(f"[أزرق] {actor}: {ef}")
        return "\n".join(out) if out else "—"

    # Always emit one row per component for this phase, even if empty
    for comp_name, arr_key in COMPONENTS:
        arr = snap.get(arr_key, []) or []
        red_entries  = entries_for_side(arr, "R")
        blue_entries = entries_for_side(arr, "B")

        # Per-component inventory + losses + ammo (this step)
        red_units_here  = units_for_component(red_snap_units,  comp_name)
        blue_units_here = units_for_component(blue_snap_units, comp_name)

        rows_for_phase.append({
            "comp":          comp_name,
            "red":           actions_list(red_entries),
            "red_why":       list_block(red_entries, fmt_why),
            "blue":          actions_list(blue_entries),
            "blue_why":      list_block(blue_entries, fmt_why),
            "combined_effect": combined_effects(red_entries, blue_entries),
            "red_inv":       inventory_str(red_units_here),
            "red_cum_loss":  cum_losses_for_comp(STATE, comp_name, "RED",  step_idx),
            "blue_inv":      inventory_str(blue_units_here),
            "blue_cum_loss": cum_losses_for_comp(STATE, comp_name, "BLUE", step_idx),
            "red_ammo":      ammo_str(red_units_here),
            "blue_ammo":     ammo_str(blue_units_here),
        })

    # Tinted fills
    RED_BLOCK   = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")
    BLUE_BLOCK  = PatternFill(start_color="eff6ff", end_color="eff6ff", fill_type="solid")
    EFFECT_FILL = PatternFill(start_color="fefce8", end_color="fefce8", fill_type="solid")  # pale yellow
    RED_LOG     = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")
    BLUE_LOG    = PatternFill(start_color="bfdbfe", end_color="bfdbfe", fill_type="solid")
    RED_AMMO    = PatternFill(start_color="fb923c", end_color="fb923c", fill_type="solid")
    BLUE_AMMO   = PatternFill(start_color="93c5fd", end_color="93c5fd", fill_type="solid")

    start = current_row
    for row_data in rows_for_phase:
        ws.cell(row=current_row, column=1,  value=phase_label)
        ws.cell(row=current_row, column=2,  value=row_data["comp"])
        ws.cell(row=current_row, column=3,  value=row_data["red"])
        ws.cell(row=current_row, column=4,  value=row_data["red_why"])
        ws.cell(row=current_row, column=5,  value=row_data["blue"])
        ws.cell(row=current_row, column=6,  value=row_data["blue_why"])
        ws.cell(row=current_row, column=7,  value=row_data["combined_effect"])
        ws.cell(row=current_row, column=8,  value=row_data["red_inv"])
        ws.cell(row=current_row, column=9,  value=row_data["red_cum_loss"])
        ws.cell(row=current_row, column=10, value=row_data["blue_inv"])
        ws.cell(row=current_row, column=11, value=row_data["blue_cum_loss"])
        ws.cell(row=current_row, column=12, value=row_data["red_ammo"])
        ws.cell(row=current_row, column=13, value=row_data["blue_ammo"])

        for c in range(1, 14):
            cell = ws.cell(row=current_row, column=c)
            cell.alignment = LEFT
            cell.border = BORDER

        # Phase col
        ws.cell(row=current_row, column=1).fill = PHASE_FILL
        ws.cell(row=current_row, column=1).font = PHASE_FONT
        ws.cell(row=current_row, column=1).alignment = CENTER

        # Component col
        comp_color = COMPONENT_COLORS.get(row_data["comp"], "f3f4f6")
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color=comp_color, end_color=comp_color, fill_type="solid")
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=2).alignment = CENTER

        # Red action + why (3-4)
        ws.cell(row=current_row, column=3).fill = RED_BLOCK
        ws.cell(row=current_row, column=4).fill = RED_BLOCK
        # Blue reaction + why (5-6)
        ws.cell(row=current_row, column=5).fill = BLUE_BLOCK
        ws.cell(row=current_row, column=6).fill = BLUE_BLOCK
        # Combined effect (7)
        ws.cell(row=current_row, column=7).fill = EFFECT_FILL
        ws.cell(row=current_row, column=7).font = Font(bold=True)
        # Logistics
        ws.cell(row=current_row, column=8).fill = RED_LOG
        ws.cell(row=current_row, column=9).fill = RED_LOG
        ws.cell(row=current_row, column=10).fill = BLUE_LOG
        ws.cell(row=current_row, column=11).fill = BLUE_LOG
        ws.cell(row=current_row, column=12).fill = RED_AMMO
        ws.cell(row=current_row, column=13).fill = BLUE_AMMO

        # Estimate row height
        line_counts = [row_data[k].count("\n") + 1 for k in
                       ("red","red_why","blue","blue_why","combined_effect",
                        "red_inv","blue_inv","red_ammo","blue_ammo")]
        ws.row_dimensions[current_row].height = max(45, min(220, 16 * max(line_counts) + 8))
        current_row += 1

    # Merge phase cell across this phase's component rows
    end = current_row - 1
    if end > start:
        ws.merge_cells(start_row=start, end_row=end, start_column=1, end_column=1)
        ws.cell(row=start, column=1).alignment = CENTER

    # Separator row
    for c in range(1, 14):
        cell = ws.cell(row=current_row, column=c)
        cell.fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        cell.border = BORDER
    ws.row_dimensions[current_row].height = 4
    current_row += 1

ws.freeze_panes = "C2"

# Save
out = ROOT / "wargameschedule.xlsx"
wb.save(out)
print(f"Saved {out} ({out.stat().st_size:,} bytes)")
print(f"Total rows used: {current_row-1}")
